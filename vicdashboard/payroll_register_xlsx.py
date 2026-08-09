"""Excel (.xlsx) builder for the payroll register."""

from __future__ import annotations

from io import BytesIO
from decimal import Decimal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .payroll_register import LEAF_COLUMNS, YELLOW_LEAF_KEYS

BLUE_FILL = PatternFill('solid', fgColor='B8D9F0')
YELLOW_FILL = PatternFill('solid', fgColor='FFF59A')
RED_FILL = PatternFill('solid', fgColor='C0392B')
LIGHT_FILL = PatternFill('solid', fgColor='F5F7FA')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')

THIN = Border(
    left=Side(style='thin', color='595959'),
    right=Side(style='thin', color='595959'),
    top=Side(style='thin', color='595959'),
    bottom=Side(style='thin', color='595959'),
)
THICK_TOP = Border(
    left=Side(style='thin', color='595959'),
    right=Side(style='thin', color='595959'),
    top=Side(style='medium', color='000000'),
    bottom=Side(style='thin', color='595959'),
)

HEADER_FONT = Font(name='Calibri', bold=True, size=8)
HEADER_FONT_WHITE = Font(name='Calibri', bold=True, size=8, color='FFFFFF')
CELL_FONT = Font(name='Calibri', size=8)
CELL_BOLD = Font(name='Calibri', bold=True, size=8)
RED_FONT = Font(name='Calibri', size=8, color='C0392B')
RED_FONT_BOLD = Font(name='Calibri', bold=True, size=8, color='C0392B')
BLUE_FONT = Font(name='Calibri', size=8, color='1A3A9C')
BLUE_FONT_BOLD = Font(name='Calibri', bold=True, size=8, color='1A3A9C')
TITLE_FONT = Font(name='Calibri', bold=True, size=14)
SIG_LABEL = Font(name='Calibri', bold=True, size=9)
SIG_NAME = Font(name='Calibri', bold=True, size=9)
SIG_TITLE = Font(name='Calibri', size=8)

CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT = Alignment(horizontal='left', vertical='center')
RIGHT = Alignment(horizontal='right', vertical='center')

MONEY_KEYS = {
    'daily_rate', 'hourly_rate', 'reg_total', 'ot_reg_amount', 'ot_sun_amount',
    'ot_total_amount', 'snwd_amount', 'snw_sun_amount', 'rh_amount', 'gross_pay',
    'philhealth', 'sss', 'hdmf', 'hdmf_loan', 'sss_loan', 'late_ut_amount',
    'total_deductions', 'net_pay',
}
RED_KEYS = {
    'philhealth', 'sss', 'hdmf', 'hdmf_loan', 'sss_loan', 'late_ut_amount', 'total_deductions',
}
INT_KEYS = {
    'no', 'reg_days', 'lwp_days', 'total_days', 'snwd_days', 'snw_sun_days', 'rh_days', 'late_ut_mins',
}


def _sheet_title(period_title: str) -> str:
    # Excel sheet name max 31 chars
    cleaned = period_title.replace('/', '-').replace('\\', '-')[:31]
    return cleaned or 'Payroll Register'


def build_payroll_register_xlsx(register: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = _sheet_title(register['meta']['period_title'])

    leaf_keys = [k for k, _ in LEAF_COLUMNS]
    leaf_labels = [label for _, label in LEAF_COLUMNS]
    n_cols = len(leaf_keys)

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(1, 1, register['meta']['period_title'])
    title_cell.font = TITLE_FONT
    title_cell.alignment = CENTER

    # Group header row (row 2)
    for label, start_key, span in register['group_headers']:
        start = leaf_keys.index(start_key) + 1
        end = start + span - 1
        if span > 1:
            ws.merge_cells(start_row=2, start_column=start, end_row=2, end_column=end)
        cell = ws.cell(2, start, label)
        cell.font = HEADER_FONT_WHITE if label == 'DEDUCTIONS' else HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
        if label == 'DEDUCTIONS':
            fill = RED_FILL
        elif label:
            fill = BLUE_FILL
        else:
            fill = WHITE_FILL
        for col in range(start, end + 1):
            c = ws.cell(2, col)
            c.fill = fill
            c.border = THIN
            c.alignment = CENTER
            if label == 'DEDUCTIONS':
                c.font = HEADER_FONT_WHITE

    # Sub-header row (row 3)
    for col, label in enumerate(leaf_labels, start=1):
        key = leaf_keys[col - 1]
        cell = ws.cell(3, col, label)
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN
        cell.fill = YELLOW_FILL if key in YELLOW_LEAF_KEYS else WHITE_FILL

    # Data rows
    data_start = 4
    for r_idx, row in enumerate(register['rows']):
        excel_row = data_start + r_idx
        for c_idx, key in enumerate(leaf_keys, start=1):
            value = row.get(key)
            cell = ws.cell(excel_row, c_idx, _excel_value(key, value))
            cell.border = THIN
            cell.font = CELL_FONT
            if key == 'name':
                cell.alignment = LEFT
            elif key in MONEY_KEYS or key.endswith('_hours') or key == 'sil_on_hand':
                cell.alignment = RIGHT
                if key in MONEY_KEYS:
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = CENTER
            if key in RED_KEYS:
                cell.font = RED_FONT
            elif key == 'net_pay':
                cell.font = BLUE_FONT

    # Totals
    totals_row = data_start + len(register['rows'])
    totals = register['totals']
    for c_idx, key in enumerate(leaf_keys, start=1):
        value = totals.get(key)
        cell = ws.cell(totals_row, c_idx, _excel_value(key, value))
        cell.border = THICK_TOP
        cell.fill = LIGHT_FILL
        cell.font = CELL_BOLD
        if key == 'name':
            cell.alignment = LEFT
        elif key in MONEY_KEYS:
            cell.alignment = RIGHT
            cell.number_format = '#,##0.00'
        else:
            cell.alignment = CENTER
        if key in RED_KEYS:
            cell.font = RED_FONT_BOLD
        elif key == 'net_pay':
            cell.font = BLUE_FONT_BOLD

    # Column widths
    for c_idx, key in enumerate(leaf_keys, start=1):
        letter = get_column_letter(c_idx)
        if key == 'name':
            ws.column_dimensions[letter].width = 22
        elif key == 'no':
            ws.column_dimensions[letter].width = 4
        else:
            ws.column_dimensions[letter].width = 9

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[3].height = 30
    ws.freeze_panes = 'A4'

    # Signatures
    sig_row = totals_row + 3
    sigs = register['signatures']
    block_width = max(1, n_cols // 4)
    for i, sig in enumerate(sigs):
        start_col = 1 + i * block_width
        end_col = start_col + block_width - 1
        if end_col > n_cols:
            end_col = n_cols
        if end_col > start_col:
            ws.merge_cells(start_row=sig_row, start_column=start_col, end_row=sig_row, end_column=end_col)
            ws.merge_cells(start_row=sig_row + 2, start_column=start_col, end_row=sig_row + 2, end_column=end_col)
            if sig.get('title'):
                ws.merge_cells(start_row=sig_row + 3, start_column=start_col, end_row=sig_row + 3, end_column=end_col)
        label_cell = ws.cell(sig_row, start_col, sig['label'])
        label_cell.font = SIG_LABEL
        label_cell.alignment = CENTER
        name_cell = ws.cell(sig_row + 2, start_col, sig['name'])
        name_cell.font = SIG_NAME
        name_cell.alignment = CENTER
        if sig.get('title'):
            title_cell = ws.cell(sig_row + 3, start_col, sig['title'])
            title_cell.font = SIG_TITLE
            title_cell.alignment = CENTER

    # Net pay callout
    callout_row = sig_row + 6
    ws.merge_cells(start_row=callout_row, start_column=max(1, n_cols - 3), end_row=callout_row, end_column=n_cols)
    callout = ws.cell(callout_row, max(1, n_cols - 3), f"TOTAL NET PAY: {_excel_value('net_pay', totals.get('net_pay'))}")
    callout.font = BLUE_FONT_BOLD
    callout.alignment = RIGHT
    if isinstance(totals.get('net_pay'), (int, float, Decimal)):
        callout.number_format = '"TOTAL NET PAY: "₱#,##0.00'

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _excel_value(key, value):
    if key == 'name' or key == 'no' and value == '':
        return value if value is not None else ''
    if value is None or value == '':
        return 0 if key != 'name' else ''
    if key in INT_KEYS:
        try:
            return int(value)
        except (TypeError, ValueError):
            return 0
    if key in MONEY_KEYS or key.endswith('_hours') or key == 'sil_on_hand':
        try:
            return float(Decimal(str(value)))
        except Exception:
            return 0.0
    return value
